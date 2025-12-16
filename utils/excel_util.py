import json
from typing import List, Dict, Any, Optional
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelUtil:
    """Excel文件操作辅助类"""
    
    @staticmethod
    def load_workbook_safe(file_path: str) -> Workbook:
        """安全加载Excel文件，增加错误处理"""

        
        path_obj = Path(file_path)
        if not path_obj.exists():
            raise FileNotFoundError(f"Excel文件不存在: {path_obj}")
        return load_workbook(file_path, data_only=True)
    
    @staticmethod
    def get_sheet(wb: Workbook, sheet_name: Optional[str] = None) -> Worksheet:
        """获取工作表，增加安全性"""
        if sheet_name and sheet_name in wb.sheetnames:
            return wb[sheet_name]
        elif wb.sheetnames:
            return wb[wb.sheetnames[0]]
        else:
            raise ValueError("Excel文件不包含任何工作表")


def read_excel(file_path: str, sheet_name: Optional[str] = None, 
              start_row: int = 2, auto_parse_json: bool = True) -> List[Dict[str, Any]]:
    """
    读取Excel文件并将其数据转换为字典列表

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称，如为None则使用第一个工作表
        start_row: 数据开始行（表头所在行+1）
        auto_parse_json: 是否自动解析JSON字符串

    Returns:
        字典列表，每个字典代表Excel文件中的一行数据

    Raises:
        FileNotFoundError: 文件不存在时
        ValueError: 文件格式不支持或工作表不存在时
    """
    wb = ExcelUtil.load_workbook_safe(file_path)
    sheet = ExcelUtil.get_sheet(wb, sheet_name)
    
    # 获取有效的数据范围
    max_row = sheet.max_row
    max_column = sheet.max_column
    
    if max_row < 1 or max_column < 1:
        return []
    
    # 读取表头
    headers = []
    for col in range(1, max_column + 1):
        header_value = sheet.cell(row=1, column=col).value
        headers.append(str(header_value) if header_value is not None else f"Column_{col}")
    
    # 读取数据行
    case_data = []
    for row in range(max(2, start_row), max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell_value = sheet.cell(row=row, column=col_idx).value
            
            # 自动解析JSON字符串
            if auto_parse_json and isinstance(cell_value, str):
                try:
                    # 尝试解析JSON
                    parsed_value = json.loads(cell_value.strip())
                    row_data[header] = parsed_value
                except (json.JSONDecodeError, AttributeError):
                    # 不是有效的JSON，保留原值
                    row_data[header] = cell_value
            else:
                row_data[header] = cell_value
        
        # 只有当行中有数据时才添加（避免空行）
        if any(value is not None for value in row_data.values()):
            case_data.append(row_data)
    
    return case_data


def write_excel(file_path: str, row_id: int, value: Any,
               target_column: Optional[str] = None, 
               sheet_name: Optional[str] = None) -> None:
    """
    将数据写入Excel文件的指定位置

    Args:
        file_path: Excel文件路径
        row_id: 行号（从1开始）
        value: 要写入的数据
        target_column: 目标列名或列标识（如"result"或"A"），如为None则自动查找"result"列
        sheet_name: 工作表名称，如为None则使用第一个工作表

    Raises:
        FileNotFoundError: 文件不存在时
        ValueError: 指定的列不存在时
    """
    wb = ExcelUtil.load_workbook_safe(file_path)
    sheet = ExcelUtil.get_sheet(wb, sheet_name)
    
    # 确定目标列
    column_idx = _resolve_target_column(sheet, target_column)
    
    # 写入数据
    sheet.cell(row=row_id, column=column_idx).value = value
    
    # 保存文件
    wb.save(file_path)


def _resolve_target_column(sheet: Worksheet, target_column: Optional[str]) -> int:
    """解析目标列的索引"""
    max_column = sheet.max_column
    
    if target_column:
        # 如果指定了列标识（如"A", "B"等）
        if target_column.isalpha():
            return ord(target_column.upper()) - ord('A') + 1
        
        # 查找匹配的列名
        for col in range(1, max_column + 1):
            header_value = sheet.cell(row=1, column=col).value
            if header_value == target_column:
                return col
        raise ValueError(f"未找到指定的列: {target_column}")
    else:
        # 默认查找"result"列
        for col in range(1, max_column + 1):
            header_value = sheet.cell(row=1, column=col).value
            if header_value and "result" in str(header_value).lower():
                return col
        
        # 如果没有找到"result"列，使用最后一列
        return max_column if max_column > 0 else 1


# 保持向后兼容的别名
read_excel_file = read_excel
write_to_excel = write_excel
