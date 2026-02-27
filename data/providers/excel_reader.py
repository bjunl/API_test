import json
from typing import Any, Optional
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from utils.logger import logger


class ExcelUtil:
    """Excel文件操作辅助类"""
    
    @staticmethod
    def load_workbook_safe(file_path: str) -> Workbook:
        """安全加载Excel文件，增加错误处理"""
        logger.debug(f"开始加载Excel文件: {file_path}")
        
        path_obj = Path(file_path)
        if not path_obj.exists():
            error_msg = f"Excel文件不存在: {path_obj}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        try:
            wb = load_workbook(file_path, data_only=True)
            logger.info(f"成功加载Excel文件: {file_path}")
            return wb
        except Exception as e:
            logger.error(f"加载Excel文件失败: {file_path}, 错误: {e}")
            raise
    
    @staticmethod
    def get_sheet(wb: Workbook, sheet_name: Optional[str] = None) -> Worksheet:
        """获取工作表，增加安全性"""
        if sheet_name and sheet_name in wb.sheetnames:
            logger.debug(f"使用指定工作表: {sheet_name}")
            return wb[sheet_name]
        elif wb.sheetnames:
            sheet_name = wb.sheetnames[0]
            logger.debug(f"使用第一个工作表: {sheet_name}")
            return wb[sheet_name]
        else:
            error_msg = "Excel文件不包含任何工作表"
            logger.error(error_msg)
            raise ValueError(error_msg)


def excel_reader(file_path: str, sheet_name: Optional[str] = None, 
              start_row: int = 2, auto_parse_json: bool = True) -> list[dict[str, Any]]:
    """
    读取Excel文件并将其数据转换为字典列表

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称，如为None则使用第一个工作表
        start_row: 数据开始行（表头所在行+1）
        auto_parse_json: 是否自动解析JSON字符串

    Returns:
        字典列表，每个字典代表Excel文件中的一行数据

    Raises:
        FileNotFoundError: 文件不存在时
        ValueError: 文件格式不支持或工作表不存在时
    """
    logger.info(f"开始读取Excel文件: {file_path}, 工作表: {sheet_name}, 起始行: {start_row}")
    
    wb = ExcelUtil.load_workbook_safe(file_path)
    sheet = ExcelUtil.get_sheet(wb, sheet_name)
    
    # 获取有效的数据范围
    max_row = sheet.max_row
    max_column = sheet.max_column
    
    logger.debug(f"Excel数据范围: 行数={max_row}, 列数={max_column}")
    
    if max_row < 1 or max_column < 1:
        logger.warning("Excel文件为空，返回空列表")
        return []
    
    # 读取表头
    headers = []
    for col in range(1, max_column + 1):
        header_value = sheet.cell(row=1, column=col).value
        headers.append(str(header_value) if header_value is not None else f"Column_{col}")
    
    logger.debug(f"读取表头: {headers}")
    
    # 读取数据行
    case_data = []
    json_parse_count = 0
    for row in range(max(2, start_row), max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell_value = sheet.cell(row=row, column=col_idx).value
            
            # 自动解析JSON字符串
            if auto_parse_json and isinstance(cell_value, str):
                try:
                    # 尝试解析JSON
                    parsed_value = json.loads(cell_value.strip())
                    row_data[header] = parsed_value
                    json_parse_count += 1
                except (json.JSONDecodeError, AttributeError):
                    # 不是有效的JSON，保留原值
                    row_data[header] = cell_value
            else:
                row_data[header] = cell_value
        
        # 只有当行中有数据时才添加（避免空行）
        if any(value is not None for value in row_data.values()):
            case_data.append(row_data)
    
    logger.info(f"Excel读取完成，共读取 {len(case_data)} 行数据，自动解析JSON {json_parse_count} 个")
    return case_data
